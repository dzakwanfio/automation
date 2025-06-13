import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    NoSuchWindowException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

# Setup logging ke file
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("otomatisasi.log")]
)

def process_files(file_paths, resume_from=2, url=None):
    # Validasi resume_from
    if not isinstance(resume_from, int) or resume_from < 1 or resume_from > 1048576:
        logging.error(f"Nilai resume_from tidak valid: {resume_from}. Harus antara 1 dan 1048576.")
        result = {
            "status": "error",
            "message": f"resume_from harus antara 1 dan 1048576. Nilai yang diberikan adalah {resume_from}",
            "last_row": 0,
        }
        print(json.dumps(result))
        return result

    # Validasi URL
    if not url:
        logging.error("URL tidak diberikan.")
        result = {
            "status": "error",
            "message": "URL destinasi harus diberikan.",
            "last_row": 0,
        }
        print(json.dumps(result))
        return result

    normalized_url = url.rstrip("/").lower()
    if normalized_url.endswith("/login"):
        normalized_url = normalized_url[:-6]  # Hapus "/login" dari akhir URL

    # Daftar situs yang diizinkan
    allowed_sites = {
        "lkp": "https://lkpdispendik.surabaya.go.id",
        "vokasi": "http://manajemen.vokasi.kemdikbud.go.id",
    }

    if normalized_url not in allowed_sites.values():
        logging.error(f"URL tidak valid: {url}")
        result = {
            "status": "error",
            "message": f"URL tidak valid: {url}. Harus https://lkpdispendik.surabaya.go.id/ atau http://manajemen.vokasi.kemdikbud.go.id/",
            "last_row": 0,
        }
        print(json.dumps(result))
        return result

    driver = None
    try:
        # Inisialisasi WebDriver Edge
        driver = webdriver.Edge()
        driver.get(url)
        logging.info(f"✅ Mengakses {url}. Silakan login di browser...")

        # Fungsi untuk mengubah nilai menjadi string aman
        def safe_str(value):
            try:
                if isinstance(value, datetime):
                    return value.strftime("%d/%m/%Y")
                return str(value).strip()
            except:
                return ""

        # Logika untuk situs Surabaya
        if normalized_url == allowed_sites["lkp"]:
            # Tunggu user login
            max_wait = 120
            wait_time = 0
            while wait_time < max_wait:
                try:
                    driver.find_element(
                        By.XPATH,
                        "//a[@href='https://lkpdispendik.surabaya.go.id/dasboard/siswa/view/list/aktif?id=966']",
                    )
                    logging.info("✅ Login terdeteksi! Melanjutkan otomatisasi...")
                    break
                except NoSuchElementException:
                    time.sleep(2)
                    wait_time += 2
                except (NoSuchWindowException, WebDriverException) as e:
                    logging.error(f"❌ Browser ditutup atau tidak dapat diakses: {e}")
                    result = {
                        "status": "error",
                        "message": "Browser ditutup sebelum login selesai",
                        "last_row": 0,
                    }
                    print(json.dumps(result))
                    return result
            else:
                logging.error("❌ Timeout: Login tidak terdeteksi dalam 2 menit.")
                result = {
                    "status": "error",
                    "message": "Timeout: Login tidak terdeteksi dalam 2 menit",
                    "last_row": 0,
                }
                print(json.dumps(result))
                return result

            # Klik link ke halaman daftar siswa aktif
            try:
                link_daftar_siswa = driver.find_element(
                    By.XPATH,
                    "//a[@href='https://lkpdispendik.surabaya.go.id/dasboard/siswa/view/list/aktif?id=966']",
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});",
                    link_daftar_siswa,
                )
                time.sleep(1)
                driver.execute_script("arguments[0].click();", link_daftar_siswa)
                logging.info("✅ Berhasil klik link 'Daftar Siswa Aktif'")
            except (NoSuchElementException, NoSuchWindowException, WebDriverException) as e:
                logging.error(f"❌ Gagal klik link 'Daftar Siswa Aktif': {e}")
                result = {
                    "status": "error",
                    "message": f"Gagal mengakses halaman daftar siswa: {str(e)}",
                    "last_row": 0,
                }
                print(json.dumps(result))
                return result

            time.sleep(3)

            # Proses semua file yang dipilih
            for file_path in file_paths:
                logging.info(f"🔵 Memulai proses file: {file_path}, exists: {os.path.exists(file_path)}")
                try:
                    wb = load_workbook(file_path)
                    sheet = wb.worksheets[0]

                    # Looping untuk membaca data dari Excel
                    i = resume_from
                    while True:
                        if sheet[f"A{i}"].value is None or str(sheet[f"A{i}"].value).strip() == "":
                            break

                        try:
                            # Membaca data dari Excel
                            nama = safe_str(sheet[f"B{i}"].value)
                            jenis_kelamin = safe_str(sheet[f"C{i}"].value)
                            NIK = safe_str(sheet[f"D{i}"].value)
                            tempat_lahir = safe_str(sheet[f"E{i}"].value)
                            tanggal_lahir = safe_str(sheet[f"F{i}"].value)
                            NISN = safe_str(sheet[f"G{i}"].value)
                            Agama = safe_str(sheet[f"H{i}"].value)
                            Handphone = safe_str(sheet[f"I{i}"].value)
                            Kewarganeraan = safe_str(sheet[f"J{i}"].value)
                            Jenis_Tinggal = safe_str(sheet[f"K{i}"].value)
                            Tanggal_Masuk = safe_str(sheet[f"L{i}"].value)
                            Email = safe_str(sheet[f"M{i}"].value)
                            Nama_Ortu = safe_str(sheet[f"N{i}"].value)
                            NIK_Ortu = safe_str(sheet[f"O{i}"].value)
                            Pekerjaan_Ortu = safe_str(sheet[f"P{i}"].value)
                            Pendidikan_Ortu = safe_str(sheet[f"Q{i}"].value)
                            Penghasilan_Ortu = safe_str(sheet[f"R{i}"].value)
                            Handphone_Ortu = safe_str(sheet[f"S{i}"].value)
                            Tempat_Lahir_Ortu = safe_str(sheet[f"T{i}"].value)
                            Tanggal_Lahir_Ortu = safe_str(sheet[f"U{i}"].value)
                            Asal = safe_str(sheet[f"V{i}"].value)
                            Alamat = safe_str(sheet[f"W{i}"].value)
                            RT = safe_str(sheet[f"X{i}"].value)
                            RW = safe_str(sheet[f"Y{i}"].value)
                            Kecamatan = safe_str(sheet[f"Z{i}"].value)
                            Kelurahan = safe_str(sheet[f"AA{i}"].value)
                            Kab_Kota = safe_str(sheet[f"AB{i}"].value)
                            Propinsi = safe_str(sheet[f"AC{i}"].value)
                            Nama_Ibu_kandung = safe_str(sheet[f"AD{i}"].value)
                            Nama_Ayah = safe_str(sheet[f"AE{i}"].value)
                            Agama_Kemdikbud = safe_str(sheet[f"AF{i}"].value)
                            Penerima_KPS = safe_str(sheet[f"AG{i}"].value)
                            Layak_PIP = safe_str(sheet[f"AH{i}"].value)
                            Penerima_KIP = safe_str(sheet[f"AI{i}"].value)
                            Kode_Pos = safe_str(sheet[f"AJ{i}"].value)
                            Jenis_tinggal = safe_str(sheet[f"AK{i}"].value)
                            Alat_Transportasi = safe_str(sheet[f"AL{i}"].value)

                            logging.info(f"Memproses data baris {i}: {nama}")

                            # Klik link ke halaman tambah siswa
                            link_tambah_siswa = driver.find_element(
                                By.XPATH,
                                "//a[@href='https://lkpdispendik.surabaya.go.id/dasboard/siswa/view/tambah_siswa']",
                            )
                            driver.execute_script(
                                "arguments[0].scrollIntoView({block: 'center'});",
                                link_tambah_siswa,
                            )
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", link_tambah_siswa)

                            # Isi form tambah siswa
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located(
                                    (By.XPATH, '//*[@id="listar-content"]')
                                )
                            )
                            driver.find_element(By.NAME, "nama").send_keys(nama)
                            driver.find_element(By.NAME, "jenis_kelamin").send_keys(jenis_kelamin)
                            driver.find_element(By.NAME, "nik").send_keys(NIK)
                            driver.find_element(By.NAME, "tempat_lahir").send_keys(tempat_lahir)
                            driver.find_element(By.NAME, "tgl_lahir").send_keys(tanggal_lahir)
                            driver.find_element(By.NAME, "nisn").send_keys(NISN)
                            driver.find_element(By.NAME, "agama").send_keys(Agama)
                            driver.find_element(By.NAME, "hp").send_keys(Handphone)
                            driver.find_element(By.NAME, "kewarganegaraan").send_keys(Kewarganeraan)
                            driver.find_element(By.NAME, "jenis_tinggal").send_keys(Jenis_Tinggal)
                            driver.find_element(By.NAME, "tgl_masuk").send_keys(Tanggal_Masuk)
                            driver.find_element(By.NAME, "email").send_keys(Email)
                            driver.find_element(By.NAME, "nama_ibu").send_keys(
                                Nama_Ibu_kandung or Nama_Ortu
                            )
                            driver.find_element(By.NAME, "nik_ibu").send_keys(NIK_Ortu)
                            driver.find_element(By.NAME, "pekerjaan_ibu").send_keys(Pekerjaan_Ortu)
                            driver.find_element(By.NAME, "pendidikan_ibu").send_keys(Pendidikan_Ortu)
                            driver.find_element(By.NAME, "penghasilan_ibu").send_keys(Penghasilan_Ortu)
                            driver.find_element(By.NAME, "hp_ibu").send_keys(Handphone_Ortu)
                            driver.find_element(By.NAME, "tempat_lahir_ibu").send_keys(Tempat_Lahir_Ortu)
                            driver.find_element(By.NAME, "tgl_lahir_ibu").send_keys(Tanggal_Lahir_Ortu)
                            driver.find_element(By.NAME, "nama_ayah").send_keys(Nama_Ayah or Nama_Ortu)
                            driver.find_element(By.NAME, "nik_ayah").send_keys(NIK_Ortu)
                            driver.find_element(By.NAME, "pekerjaan_ayah").send_keys(Pekerjaan_Ortu)
                            driver.find_element(By.NAME, "pendidikan_ayah").send_keys(Pendidikan_Ortu)
                            driver.find_element(By.NAME, "penghasilan_ayah").send_keys(Penghasilan_Ortu)
                            driver.find_element(By.NAME, "hp_ayah").send_keys(Handphone_Ortu)
                            driver.find_element(By.NAME, "tempat_lahir_ayah").send_keys(Tempat_Lahir_Ortu)
                            driver.find_element(By.NAME, "tgl_lahir_ayah").send_keys(Tanggal_Lahir_Ortu)
                            driver.find_element(By.NAME, "asal_domisili").send_keys("SURABAYA" if safe_str(sheet[f"V{i}"].value).strip().lower() == "surabaya" else "Luar Kota")
                            driver.find_element(By.NAME, "alamat_domisili").send_keys(Alamat)
                            driver.find_element(By.NAME, "rt_domisili").send_keys(RT)
                            driver.find_element(By.NAME, "rw_domisili").send_keys(RW)
                            driver.find_element(By.NAME, "kecamatan_domisili").send_keys(Kecamatan)
                            driver.find_element(By.NAME, "kelurahan_domisili").send_keys(Kelurahan)
                            driver.find_element(By.NAME, "asal_kk").send_keys(Asal)
                            driver.find_element(By.NAME, "alamat_kk").send_keys(Alamat)
                            driver.find_element(By.NAME, "rt_kk").send_keys(RT)
                            driver.find_element(By.NAME, "rw_kk").send_keys(RW)
                            driver.find_element(By.NAME, "kecamatan_kk").send_keys(Kecamatan)
                            driver.find_element(By.NAME, "kelurahan_kk").send_keys(Kelurahan)

                            driver.find_element(By.XPATH, "//button[@type='submit']").click()
                            logging.info(f"✅ Baris {i} berhasil diproses")
                        except (NoSuchWindowException, WebDriverException) as e:
                            logging.error(f"❌ Browser ditutup pada baris {i}: {e}")
                            result = {
                                "status": "error",
                                "message": f"Terdeteksi berhenti pada baris {i}",
                                "last_row": i,
                            }
                            print(json.dumps(result))
                            return result
                        except (NoSuchElementException, TimeoutException) as e:
                            logging.error(f"❌ Gagal memproses baris {i}: {e}")
                            result = {
                                "status": "error",
                                "message": f"Gagal memproses baris {i}: {str(e)}",
                                "last_row": i,
                            }
                            print(json.dumps(result))
                            return result
                        except Exception as e:
                            logging.error(f"❌ Error tidak terduga pada baris {i}: {e}")
                            result = {
                                "status": "error",
                                "message": f"Error tidak terduga pada baris {i}: {str(e)}",
                                "last_row": i,
                            }
                            print(json.dumps(result))
                            return result

                        i += 1

                    logging.info(f"✅ File {file_path} selesai diproses, exists: {os.path.exists(file_path)}")

                except (NoSuchWindowException, WebDriverException) as e:
                    logging.error(f"❌ Browser ditutup saat memproses file {file_path}: {e}")
                    result = {
                        "status": "error",
                        "message": f"Terdeteksi berhenti pada baris {i}",
                        "last_row": i,
                    }
                    print(json.dumps(result))
                    return result
                except Exception as e:
                    logging.error(f"❌ Error saat memproses file {file_path}: {e}")
                    result = {
                        "status": "error",
                        "message": f"Error memproses file {file_path}: {str(e)}",
                        "last_row": i,
                    }
                    print(json.dumps(result))
                    return result

        # Logika untuk situs Vokasi
        elif normalized_url == allowed_sites["vokasi"]:
            # Tunggu user login
            max_wait = 120
            wait_time = 0
            while wait_time < max_wait:
                try:
                    driver.find_element(
                        By.XPATH, "//a[contains(text(),'Peserta Didik')]"
                    )
                    logging.info("✅ Login terdeteksi! Melanjutkan otomatisasi...")
                    break
                except NoSuchElementException:
                    time.sleep(2)
                    wait_time += 2
                except (NoSuchWindowException, WebDriverException) as e:
                    logging.error(f"❌ Browser ditutup atau tidak dapat diakses: {e}")
                    result = {
                        "status": "error",
                        "message": "Browser ditutup sebelum login selesai",
                        "last_row": 0,
                    }
                    print(json.dumps(result))
                    return result
            else:
                logging.error("❌ Timeout: Login tidak terdeteksi dalam 2 menit.")
                result = {
                    "status": "error",
                    "message": "Timeout: Login tidak terdeteksi dalam 2 menit",
                    "last_row": 0,
                }
                print(json.dumps(result))
                return result

            time.sleep(5)

            # Klik link ke halaman "Data PD"
            try:
                link_daftar_siswa = driver.find_element(
                    By.XPATH, "//a[contains(text(),'Data PD')]"
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});",
                    link_daftar_siswa,
                )
                time.sleep(1)
                driver.execute_script("arguments[0].click();", link_daftar_siswa)
                logging.info("✅ Berhasil klik link 'Data PD'")
            except Exception as e:
                logging.error(f"❌ Gagal klik link 'Data PD': {e}")
                result = {
                    "status": "error",
                    "message": f"Gagal mengakses Data PD: {str(e)}",
                    "last_row": 0,
                }
                print(json.dumps(result))
                return result

            time.sleep(3)

            # Proses semua file yang dipilih
            for file_path in file_paths:
                logging.info(f"🔵 Memulai proses file: {file_path}, exists: {os.path.exists(file_path)}")
                try:
                    wb = load_workbook(file_path)
                    sheet = wb.worksheets[0]

                    # Looping untuk membaca data dari Excel
                    i = resume_from
                    while True:
                        if sheet[f"A{i}"].value is None or str(sheet[f"A{i}"].value).strip() == "":
                            logging.info(f"Kolom A pada baris {i} kosong. Loop berhenti.")
                            break

                        try:
                            # Membaca data dari Excel
                            nama = safe_str(sheet[f"B{i}"].value)
                            jenis_kelamin = safe_str(sheet[f"C{i}"].value)
                            NIK = safe_str(sheet[f"D{i}"].value)
                            tempat_lahir = safe_str(sheet[f"E{i}"].value)
                            Tanggal_Lahir = safe_str(sheet[f"F{i}"].value)
                            NISN = safe_str(sheet[f"G{i}"].value)
                            Agama_LKP = safe_str(sheet[f"H{i}"].value)
                            Handphone = safe_str(sheet[f"I{i}"].value)
                            Kewarganeraan = safe_str(sheet[f"J{i}"].value)
                            Jenis_Tinggal = safe_str(sheet[f"K{i}"].value)
                            Tanggal_Masuk = safe_str(sheet[f"L{i}"].value)
                            Email = safe_str(sheet[f"M{i}"].value)
                            Nama_Ibu_kandung = safe_str(sheet[f"AD{i}"].value)
                            NIK_Ibu = safe_str(sheet[f"O{i}"].value)
                            Pekerjaan_Ibu = safe_str(sheet[f"P{i}"].value)
                            Pendidikan_Ibu = safe_str(sheet[f"Q{i}"].value)
                            Penghasilan_Ibu = safe_str(sheet[f"R{i}"].value)
                            Handphone_Ibu = safe_str(sheet[f"S{i}"].value)
                            Tempat_Lahir_Ibu = safe_str(sheet[f"T{i}"].value)
                            Tanggal_Lahir_Ibu = safe_str(sheet[f"U{i}"].value)
                            Nama_Ayah = safe_str(sheet[f"AE{i}"].value)
                            NIK_Ayah = safe_str(sheet[f"O{i}"].value)
                            Pekerjaan_Ayah = safe_str(sheet[f"P{i}"].value)
                            Pendidikan_Ayah = safe_str(sheet[f"Q{i}"].value)
                            Penghasilan_Ayah = safe_str(sheet[f"R{i}"].value)
                            Handphone_Ayah = safe_str(sheet[f"S{i}"].value)
                            Tempat_Lahir_Ayah = safe_str(sheet[f"T{i}"].value)
                            Tanggal_Lahir_Ayah = safe_str(sheet[f"U{i}"].value)
                            Asal_Domisili = safe_str(sheet[f"V{i}"].value)
                            Alamat_Domisili = safe_str(sheet[f"W{i}"].value)
                            RT_Domisili = safe_str(sheet[f"X{i}"].value)
                            RW_Domisili = safe_str(sheet[f"Y{i}"].value)
                            Kecamatan_Domisili = safe_str(sheet[f"Z{i}"].value)
                            Kelurahan_Domisili = safe_str(sheet[f"AA{i}"].value).lower()
                            Kota = safe_str(sheet[f"AB{i}"].value)
                            Provinsi = safe_str(sheet[f"AC{i}"].value)
                            Agama_Kemdikbud = safe_str(sheet[f"AF{i}"].value)
                            Penerima_KPS = safe_str(sheet[f"AG{i}"].value)
                            Layak_PIP = safe_str(sheet[f"AH{i}"].value)
                            Penerima_KIP = safe_str(sheet[f"AI{i}"].value)
                            Kode_Pos = safe_str(sheet[f"AJ{i}"].value)
                            Jenis_Tinggal_Form = safe_str(sheet[f"AK{i}"].value)
                            Transportasi = safe_str(sheet[f"AL{i}"].value)

                            logging.info(f"Memproses data baris {i}: {nama}")

                            # Klik tombol "Peserta Didik Baru"
                            try:
                                link_tambah_siswa = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, "//button[contains(., 'Peserta Didik Baru')]")
                                    )
                                )
                                driver.execute_script(
                                    "arguments[0].scrollIntoView({block: 'center'});",
                                    link_tambah_siswa,
                                )
                                time.sleep(1)
                                driver.execute_script("arguments[0].click();", link_tambah_siswa)
                                logging.info("✅ Pop-up 'Peserta Didik Baru' terbuka.")
                            except Exception as e:
                                logging.error(f"❌ Gagal membuka pop-up 'Peserta Didik Baru': {e}")
                                result = {
                                    "status": "error",
                                    "message": f"Gagal membuka pop-up 'Peserta Didik Baru': {str(e)}",
                                    "last_row": i,
                                }
                                print(json.dumps(result))
                                return result

                            # Pastikan pop-up terbuka
                            retry_attempts = 3
                            popup_opened = False
                            for attempt in range(retry_attempts):
                                try:
                                    WebDriverWait(driver, 15).until(
                                        EC.presence_of_element_located((By.ID, "nama"))
                                    )
                                    popup_opened = True
                                    break
                                except TimeoutException:
                                    logging.warning(
                                        f"Percobaan {attempt+1}/{retry_attempts}: Pop-up belum terbuka. Mencoba lagi."
                                    )
                                    try:
                                        link_tambah_siswa = WebDriverWait(driver, 15).until(
                                            EC.presence_of_element_located(
                                                (By.XPATH, "//button[contains(., 'Peserta Didik Baru')]")
                                            )
                                        )
                                        driver.execute_script(
                                            "arguments[0].scrollIntoView({block: 'center'});",
                                            link_tambah_siswa,
                                        )
                                        time.sleep(1)
                                        driver.execute_script("arguments[0].click();", link_tambah_siswa)
                                        logging.info(
                                            f"Percobaan {attempt+1}/{retry_attempts}: Pop-up dibuka kembali."
                                        )
                                    except TimeoutException:
                                        logging.error(
                                            f"Percobaan {attempt+1}/{retry_attempts}: Gagal menemukan tombol."
                                        )
                                        continue

                            if not popup_opened:
                                logging.error(
                                    f"Gagal membuka pop-up setelah {retry_attempts} percobaan."
                                )
                                driver.get(
                                    "http://manajemen.vokasi.kemdikbud.go.id/peserta-didik/data-pd"
                                )
                                time.sleep(5)
                                try:
                                    link_tambah_siswa = WebDriverWait(driver, 15).until(
                                        EC.presence_of_element_located(
                                            (By.XPATH, "//button[contains(., 'Peserta Didik Baru')]")
                                        )
                                    )
                                    driver.execute_script(
                                        "arguments[0].scrollIntoView({block: 'center'});",
                                        link_tambah_siswa,
                                    )
                                    time.sleep(1)
                                    driver.execute_script("arguments[0].click();", link_tambah_siswa)
                                    logging.info("Pop-up dibuka kembali setelah navigasi ulang.")
                                except Exception as e:
                                    logging.error(f"Gagal navigasi ulang: {str(e)}")
                                    result = {
                                        "status": "error",
                                        "message": f"Gagal navigasi ulang: {str(e)}",
                                        "last_row": i,
                                    }
                                    print(json.dumps(result))
                                    return result

                            # Isi form pencarian
                            try:
                                nama_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "nama"))
                                )
                                driver.execute_script(
                                    "arguments[0].scrollIntoView({block: 'center'});",
                                    nama_field,
                                )
                                nama_field.clear()
                                nama_field.send_keys(nama)

                                nik_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "nik"))
                                )
                                nik_field.clear()
                                nik_field.send_keys(NIK)

                                nisn_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "nisn"))
                                )
                                nisn_field.clear()
                                nisn_field.send_keys(NISN)

                                tempat_lahir_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "tempat_lahir"))
                                )
                                tempat_lahir_field.clear()
                                tempat_lahir_field.send_keys(tempat_lahir)

                                tanggal_lahir_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "tanggal_lahir"))
                                )
                                driver.execute_script(
                                    "arguments[0].scrollIntoView({block: 'center'});",
                                    tanggal_lahir_field,
                                )
                                tanggal_lahir_field.clear()
                                tanggal_lahir_field.send_keys(Tanggal_Lahir)
                                logging.info(f"Tanggal Lahir diisi: {Tanggal_Lahir}")

                                nama_ibu_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "nama_ibu_kandung"))
                                )
                                nama_ibu_field.clear()
                                nama_ibu_field.send_keys(Nama_Ibu_kandung)

                                logging.info(
                                    f"Formulir pencarian diisi untuk baris {i}: Nama={nama}, NIK={NIK}, NISN={NISN}, Nama Ibu={Nama_Ibu_kandung}"
                                )

                                # Klik tombol "Cari"
                                tombol_cari = WebDriverWait(driver, 20).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, "//button[contains(text(), 'Cari')]")
                                    )
                                )
                                tombol_cari.click()
                                logging.info(f"Klik tombol Cari berhasil untuk baris {i}")
                                time.sleep(3)
                            except Exception as e:
                                logging.error(f"Gagal mengisi form pencarian untuk baris {i}: {str(e)}")
                                driver.save_screenshot(f"error_form_row_{i}.png")
                                i += 1
                                continue

                            # Cek apakah data tidak ditemukan
                            try:
                                WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable(
                                        (
                                            By.CSS_SELECTOR,
                                            "input[wire\\:model\\.lazy='pd_nama_ayah']",
                                        )
                                    )
                                )
                                logging.info(
                                    f"Data tidak ditemukan untuk baris {i}. Mengisi formulir manual"
                                )

                                # Isi form manual
                                nama_ayah_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable(
                                        (
                                            By.CSS_SELECTOR,
                                            "input[wire\\:model\\.lazy='pd_nama_ayah']",
                                        )
                                    )
                                )
                                nama_ayah_field.clear()
                                nama_ayah_field.send_keys(Nama_Ayah)

                                pd_nisn_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-nisn"))
                                )
                                pd_nisn_field.clear()
                                pd_nisn_field.send_keys(NISN)

                                pd_nik_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-nik"))
                                )
                                pd_nik_field.clear()
                                pd_nik_field.send_keys(NIK)

                                jenis_kelamin_value = {"l": "L", "p": "P"}.get(
                                    jenis_kelamin.lower(), "L"
                                )
                                Select(
                                    driver.find_element(By.ID, "pd-jenis-kelamin")
                                ).select_by_value(jenis_kelamin_value)
                                logging.info(f"Jenis kelamin diisi: {jenis_kelamin_value}")

                                Select(driver.find_element(By.ID, "pd-agama")).select_by_value(
                                    {
                                        "islam": "1",
                                        "kristen": "2",
                                        "katholik": "3",
                                        "hindu": "4",
                                        "budha": "5",
                                        "khonghucu": "6",
                                        "kepercayaan": "7",
                                        "kepercayaan kpd tuhan yme": "7",
                                        "lainnya": "99",
                                    }.get(Agama_Kemdikbud.lower(), "1")
                                )

                                Select(
                                    driver.find_element(By.ID, "pd-kewarganegaraan")
                                ).select_by_value(
                                    {
                                        "afghanistan": "AF",
                                        "aland islands": "AX",
                                        "albania": "AL",
                                        "algeria": "DZ",
                                        "american samoa": "AS",
                                        "andorra": "AD",
                                        "angola": "AO",
                                        "anguilla": "AI",
                                        "antigua and barbuda": "AG",
                                        "argentina": "AR",
                                        "armenia": "AM",
                                        "aruba": "AW",
                                        "australia": "AU",
                                        "austria": "AT",
                                        "azerbaijan": "AZ",
                                        "bahamas": "BS",
                                        "bahrain": "BH",
                                        "bangladesh": "BD",
                                        "barbados": "BB",
                                        "belarus": "BY",
                                        "belgium": "BE",
                                        "belize": "BZ",
                                        "benin": "BJ",
                                        "bermuda": "BM",
                                        "bhutan": "BT",
                                        "bolivia": "BO",
                                        "bosnia and herzegovina": "BA",
                                        "botswana": "BW",
                                        "brazil": "BR",
                                        "british indian ocean territory": "IO",
                                        "brunei darussalam": "BN",
                                        "bulgaria": "BG",
                                        "burkina faso": "BF",
                                        "burundi": "BI",
                                        "cabo verde": "CV",
                                        "cambodia": "KH",
                                        "cameroon": "CM",
                                        "canada": "CA",
                                        "cayman islands": "KY",
                                        "central african republic": "CF",
                                        "chad": "TD",
                                        "chile": "CL",
                                        "china": "CN",
                                        "colombia": "CO",
                                        "comoros": "KM",
                                        "congo": "CG",
                                        "congo (democratic republic)": "CD",
                                        "cook islands": "CK",
                                        "costa rica": "CR",
                                        "croatia": "HR",
                                        "cuba": "CU",
                                        "cyprus": "CY",
                                        "czech republic": "CZ",
                                        "denmark": "DK",
                                        "djibouti": "DJ",
                                        "dominica": "DM",
                                        "dominican republic": "DO",
                                        "ecuador": "EC",
                                        "egypt": "EG",
                                        "el salvador": "SV",
                                        "equatorial guinea": "GQ",
                                        "eritrea": "ER",
                                        "estonia": "EE",
                                        "eswatini": "SZ",
                                        "ethiopia": "ET",
                                        "fiji": "FJ",
                                        "finland": "FI",
                                        "france": "FR",
                                        "gabon": "GA",
                                        "gambia": "GM",
                                        "georgia": "GE",
                                        "germany": "DE",
                                        "ghana": "GH",
                                        "greece": "GR",
                                        "grenada": "GD",
                                        "guam": "GU",
                                        "guatemala": "GT",
                                        "guinea": "GN",
                                        "guinea-bissau": "GW",
                                        "guyana": "GY",
                                        "haiti": "HT",
                                        "honduras": "HN",
                                        "hong kong": "HK",
                                        "hungary": "HU",
                                        "iceland": "IS",
                                        "india": "IN",
                                        "indonesia": "ID",
                                        "iran": "IR",
                                        "iraq": "IQ",
                                        "ireland": "IE",
                                        "israel": "IL",
                                        "italy": "IT",
                                        "jamaica": "JM",
                                        "japan": "JP",
                                        "jordan": "JO",
                                        "kazakhstan": "KZ",
                                        "kenya": "KE",
                                        "kiribati": "KI",
                                        "kosovo": "XK",
                                        "kuwait": "KW",
                                        "kyrgyzstan": "KG",
                                        "lao people's democratic republic": "LA",
                                        "latvia": "LV",
                                        "lebanon": "LB",
                                        "lesotho": "LS",
                                        "liberia": "LR",
                                        "libya": "LY",
                                        "liechtenstein": "LI",
                                        "lithuania": "LT",
                                        "luxembourg": "LU",
                                        "madagascar": "MG",
                                        "malawi": "MW",
                                        "malaysia": "MY",
                                        "maldives": "MV",
                                        "mali": "ML",
                                        "malta": "MT",
                                        "marshall islands": "MH",
                                        "mauritania": "MR",
                                        "mauritius": "MU",
                                        "mexico": "MX",
                                        "micronesia (federated states of)": "FM",
                                        "moldova": "MD",
                                        "monaco": "MC",
                                        "mongolia": "MN",
                                        "montserrat": "MS",
                                        "morocco": "MA",
                                        "mozambique": "MZ",
                                        "myanmar": "MM",
                                        "namibia": "NA",
                                        "nauru": "NR",
                                        "nepal": "NP",
                                        "netherlands": "NL",
                                        "new zealand": "NZ",
                                        "nicaragua": "NI",
                                        "niger": "NE",
                                        "nigeria": "NG",
                                        "niue": "NU",
                                        "norfolk island": "NF",
                                        "north korea": "KP",
                                        "north macedonia": "MK",
                                        "northern mariana islands": "MP",
                                        "norway": "NO",
                                        "oman": "OM",
                                        "pakistan": "PK",
                                        "palau": "PW",
                                        "panama": "PA",
                                        "papua new guinea": "PG",
                                        "paraguay": "PY",
                                        "peru": "PE",
                                        "philippines": "PH",
                                        "pitcairn": "PN",
                                        "poland": "PL",
                                        "portugal": "PT",
                                        "puerto rico": "PR",
                                        "qatar": "QA",
                                        "reunion": "RE",
                                        "romania": "RO",
                                        "russia": "RU",
                                        "rwanda": "RW",
                                        "saint barthelemy": "BL",
                                        "saint helena, ascension and tristan da cunha": "SH",
                                        "saint kitts and nevis": "KN",
                                        "saint lucia": "LC",
                                        "saint martin": "MF",
                                        "saint pierre and miquelon": "PM",
                                        "saint vincent and the grenadines": "VC",
                                        "samoa": "WS",
                                        "san marino": "SM",
                                        "sao tome and principe": "ST",
                                        "saudi arabia": "SA",
                                        "senegal": "SN",
                                        "serbia": "RS",
                                        "seychelles": "SC",
                                        "sierra leone": "SL",
                                        "singapore": "SG",
                                        "slovakia": "SK",
                                        "slovenia": "SI",
                                        "solomon islands": "SB",
                                        "somalia": "SO",
                                        "south africa": "ZA",
                                        "south georgia and the south sandwich islands": "GS",
                                        "south korea": "KR",
                                        "south sudan": "SS",
                                        "spain": "ES",
                                        "sri lanka": "LK",
                                        "sudan": "SD",
                                        "suriname": "SR",
                                        "svalbard and jan mayen": "SJ",
                                        "sweden": "SE",
                                        "switzerland": "CH",
                                        "syrian arab republic": "SY",
                                        "taiwan": "TW",
                                        "tajikistan": "TJ",
                                        "tanzania": "TZ",
                                        "thailand": "TH",
                                        "timor-leste": "TL",
                                        "togo": "TG",
                                        "tokelau": "TK",
                                        "tonga": "TO",
                                        "trinidad and tobago": "TT",
                                        "tunisia": "TN",
                                        "turkmenistan": "TM",
                                        "turkey": "TR",
                                        "tuvalu": "TV",
                                        "uganda": "UG",
                                        "ukraine": "UA",
                                        "united arab emirates": "AE",
                                        "united kingdom": "GB",
                                        "united states of america": "US",
                                        "uruguay": "UY",
                                        "uzbekistan": "UZ",
                                        "vanuatu": "VU",
                                        "venezuela": "VE",
                                        "vietnam": "VN",
                                        "western sahara": "EH",
                                        "yemen": "YE",
                                        "zambia": "ZM",
                                        "zimbabwe": "ZW",
                                    }.get(Kewarganeraan.lower(), "ID")
                                )

                                alamat_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-alamat"))
                                )
                                alamat_field.clear()
                                alamat_field.send_keys(Alamat_Domisili)

                                rt_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-rt"))
                                )
                                rt_field.clear()
                                rt_field.send_keys(RT_Domisili)

                                rw_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-rw"))
                                )
                                rw_field.clear()
                                rw_field.send_keys(RW_Domisili)

                                kodepos_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-kodepos"))
                                )
                                kodepos_field.clear()
                                kodepos_field.send_keys(Kode_Pos)

                                dusun_field = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable((By.ID, "pd-dusun"))
                                )
                                dusun_field.clear()
                                dusun_field.send_keys(Kelurahan_Domisili)

                                Select(driver.find_element(By.ID, "pd-provinsi")).select_by_value(
                                    {
                                        "luar negeri": "350000  ",
                                        "aceh": "060000  ",
                                        "bali": "220000  ",
                                        "banten": "280000  ",
                                        "bengkulu": "260000  ",
                                        "d.i. yogyakarta": "040000  ",
                                        "yogyakarta": "040000  ",
                                        "d.k.i. jakarta": "010000  ",
                                        "jakarta": "010000  ",
                                        "gorontalo": "300000  ",
                                        "jambi": "100000  ",
                                        "jawa barat": "020000  ",
                                        "jawa tengah": "030000  ",
                                        "jawa timur": "050000  ",
                                        "kalimantan barat": "130000  ",
                                        "kalimantan selatan": "150000  ",
                                        "kalimantan tengah": "140000  ",
                                        "kalimantan timur": "160000  ",
                                        "kalimantan utara": "340000  ",
                                        "kepulauan bangka belitung": "290000  ",
                                        "kepulauan riau": "310000  ",
                                        "lampung": "120000  ",
                                        "maluku": "210000  ",
                                        "maluku utara": "270000  ",
                                        "nusa tenggara barat": "230000  ",
                                        "nusa tenggara timur": "240000  ",
                                        "papua": "250000  ",
                                        "papua barat": "320000  ",
                                        "papua barat daya": "390000  ",
                                        "papua pegunungan": "380000  ",
                                        "papua selatan": "370000  ",
                                        "papua tengah": "360000  ",
                                        "riau": "090000  ",
                                        "sulawesi barat": "330000  ",
                                        "sulawesi selatan": "190000  ",
                                        "sulawesi tengah": "180000  ",
                                        "sulawesi tenggara": "200000  ",
                                        "sulawesi utara": "170000  ",
                                        "sumatera barat": "080000  ",
                                        "sumatera selatan": "110000  ",
                                        "sumatera utara": "070000  ",
                                    }.get(Provinsi.lower(), "010000  ")
                                )

                                WebDriverWait(driver, 20).until(
                                    lambda d: len(
                                        Select(d.find_element(By.ID, "pd-kabupatenkota")).options
                                    )
                                    > 1
                                )
                                Select(
                                    driver.find_element(By.ID, "pd-kabupatenkota")
                                ).select_by_visible_text(Kota)

                                WebDriverWait(driver, 20).until(
                                    lambda d: len(
                                        Select(d.find_element(By.ID, "pd-kecamatan")).options
                                    )
                                    > 1
                                )
                                Select(
                                    driver.find_element(By.ID, "pd-kecamatan")
                                ).select_by_visible_text(Kecamatan_Domisili)

                                WebDriverWait(driver, 20).until(
                                    lambda d: len(
                                        Select(d.find_element(By.ID, "pd-kelurahan")).options
                                    )
                                    > 1
                                )
                                for option in Select(
                                    driver.find_element(By.ID, "pd-kelurahan")
                                ).options:
                                    if option.text.strip().lower() == Kelurahan_Domisili:
                                        option.click()
                                        break

                                Select(
                                    driver.find_element(By.ID, "pd-jenis-tinggal")
                                ).select_by_visible_text(Jenis_Tinggal_Form)

                                Select(
                                    driver.find_element(By.ID, "pd-transportasi")
                                ).select_by_visible_text(Transportasi)

                                Select(
                                    driver.find_element(By.ID, "pd-penerima-kps")
                                ).select_by_value(
                                    {"ya": "1", "tidak": "0"}.get(Penerima_KPS.lower(), "0")
                                )
                                Select(
                                    driver.find_element(By.ID, "pd-layak-pip")
                                ).select_by_value(
                                    {"layak": "1", "tidak layak": "0"}.get(
                                        Layak_PIP.lower(), "0"
                                    )
                                )
                                Select(
                                    driver.find_element(By.ID, "pd-penerima-kip")
                                ).select_by_value(
                                    {"ya": "1", "tidak": "0"}.get(Penerima_KIP.lower(), "0")
                                )

                                simpan_button = WebDriverWait(driver, 15).until(
                                    EC.element_to_be_clickable(
                                        (By.XPATH, "//button[contains(text(), 'Simpan')]")
                                    )
                                )
                                simpan_button.click()
                                logging.info(f"Formulir manual disimpan untuk baris {i}")
                            except TimeoutException:
                                logging.info(
                                    f"Data sudah ada untuk baris {i}. Melakukan refresh halaman"
                                )
                                driver.refresh()
                                time.sleep(5)
                                try:
                                    WebDriverWait(driver, 15).until(
                                        EC.presence_of_element_located(
                                            (By.XPATH, "//a[contains(text(),'Data PD')]")
                                        )
                                    )
                                    retry_attempts = 3
                                    for attempt in range(retry_attempts):
                                        try:
                                            link_tambah_siswa = WebDriverWait(driver, 15).until(
                                                EC.presence_of_element_located(
                                                    (
                                                        By.XPATH,
                                                        "//button[contains(., 'Peserta Didik Baru')]",
                                                    )
                                                )
                                            )
                                            driver.execute_script(
                                                "arguments[0].scrollIntoView({block: 'center'});",
                                                link_tambah_siswa,
                                            )
                                            time.sleep(1)
                                            driver.execute_script(
                                                "arguments[0].click();", link_tambah_siswa
                                            )
                                            logging.info(
                                                f"Pop-up 'Peserta Didik Baru' dibuka kembali untuk baris {i+1}"
                                            )
                                            nama_field = WebDriverWait(driver, 15).until(
                                                EC.element_to_be_clickable((By.ID, "nama"))
                                            )
                                            nama_field.clear()
                                            break
                                        except TimeoutException:
                                            logging.warning(
                                                f"Percobaan {attempt+1}/{retry_attempts}: Gagal menemukan tombol setelah refresh."
                                            )
                                            if attempt == retry_attempts - 1:
                                                logging.error(
                                                    f"Gagal menemukan tombol setelah {retry_attempts} percobaan untuk baris {i+1}"
                                                )
                                                driver.save_screenshot(
                                                    f"error_refresh_row_{i}.png"
                                                )
                                                result = {
                                                    "status": "error",
                                                    "message": f"Gagal menemukan tombol 'Peserta Didik Baru' setelah {retry_attempts} percobaan",
                                                    "last_row": i,
                                                }
                                                print(json.dumps(result))
                                                return result
                                except TimeoutException:
                                    logging.error(
                                        f"Halaman utama tidak dimuat dengan benar setelah refresh untuk baris {i+1}"
                                    )
                                    driver.save_screenshot(f"error_refresh_row_{i}.png")
                                    result = {
                                        "status": "error",
                                        "message": f"Halaman utama tidak dimuat dengan benar setelah refresh",
                                        "last_row": i,
                                    }
                                    print(json.dumps(result))
                                    return result
                                i += 1
                                continue
                            except Exception as e:
                                logging.error(
                                    f"Gagal mengisi form manual untuk baris {i}: {str(e)}"
                                )
                                driver.save_screenshot(f"error_manual_form_row_{i}.png")
                                i += 1
                                continue

                            i += 1

                        except (NoSuchWindowException, WebDriverException) as e:
                            logging.error(f"Browser ditutup pada baris {i}: {e}")
                            result = {
                                "status": "error",
                                "message": f"Terdeteksi berhenti pada baris {i}",
                                "last_row": i,
                            }
                            print(json.dumps(result))
                            return result
                        except Exception as e:
                            logging.error(f"Error tidak terduga pada baris {i}: {e}")
                            result = {
                                "status": "error",
                                "message": f"Error tidak terduga pada baris {i}: {str(e)}",
                                "last_row": i,
                            }
                            print(json.dumps(result))
                            return result

                    logging.info(f"✅ File {file_path} selesai diproses, exists: {os.path.exists(file_path)}")

                except (NoSuchWindowException, WebDriverException) as e:
                    logging.error(f"Browser ditutup saat memproses file {file_path}: {e}")
                    result = {
                        "status": "error",
                        "message": f"Terdeteksi berhenti pada baris {i}",
                        "last_row": i,
                    }
                    print(json.dumps(result))
                    return result
                except Exception as e:
                    logging.error(f"Error saat memproses file {file_path}: {e}")
                    result = {
                        "status": "error",
                        "message": f"Error memproses file {file_path}: {str(e)}",
                        "last_row": i,
                    }
                    print(json.dumps(result))
                    return result

        logging.info("✅ Semua file telah diproses!")
        result = {
            "status": "success",
            "message": f"Semua {len(file_paths)} file berhasil diproses!",
            "last_row": i,
        }
        print(json.dumps(result))
        return result

    except (NoSuchWindowException, WebDriverException) as e:
        logging.error(f"❌ Browser ditutup: {e}")
        result = {
            "status": "error",
            "message": "Browser ditutup secara tidak sengaja",
            "last_row": 0,
        }
        print(json.dumps(result))
        return result
    except Exception as e:
        logging.error(f"❌ Error utama: {e}")
        result = {"status": "error", "message": str(e), "last_row": 0}
        print(json.dumps(result))
        return result
    finally:
        if driver:
            try:
                logging.info("Attempting to close all windows and quit driver...")
                for handle in driver.window_handles:
                    driver.switch_to.window(handle)
                    driver.close()
                driver.quit()
                logging.info("Driver quit successfully.")
            except Exception as e:
                logging.error(f"Failed to quit driver: {e}")

if __name__ == "__main__":
    # Parse argumen baris perintah
    parser = argparse.ArgumentParser(description="Script otomatisasi pengentry-an data siswa")
    parser.add_argument(
        "file_paths", nargs="+", help="Path ke file Excel yang akan diproses"
    )
    parser.add_argument(
        "--resume-from",
        type=int,
        default=2,
        help="Baris mulai untuk melanjutkan proses",
    )
    parser.add_argument(
        "--url",
        type=str,
        required=True,
        help="URL destinasi (lkpdispendik.surabaya.go.id atau manajemen.vokasi.kemdikbud.go.id)",
    )
    args = parser.parse_args()

    # Debugging info, dicetak ke stderr
    print("=== DEBUG INFO ===", file=sys.stderr)
    print(f"Current directory: {os.getcwd()}", file=sys.stderr)
    print(f"Script location: {os.path.abspath(__file__)}", file=sys.stderr)
    print(f"Args received: {sys.argv}", file=sys.stderr)

    if len(args.file_paths) < 1:
        print(json.dumps({"status": "error", "message": "No file paths provided", "last_row": 0}))
        sys.exit(1)

    # Verifikasi file input
    for i, file_path in enumerate(args.file_paths):
        if not os.path.exists(file_path):
            print(
                json.dumps(
                    {
                        "status": "error",
                        "message": f"File {i+1} not found - {file_path}",
                        "last_row": 0,
                    }
                )
            )
            sys.exit(1)
        print(f"File {i+1} found: {file_path}", file=sys.stderr)

    print("=================", file=sys.stderr)

    # Jalankan proses utama
    process_files(args.file_paths, resume_from=args.resume_from, url=args.url)